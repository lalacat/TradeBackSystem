import re

import openpyxl as openpyxl
import tushare as ts
import pandas as pd
import datetime as dt

from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Font
from win32com.client import Dispatch

token = 'bfbf67e56f47ef62e570fc6595d57909f9fc516d3749458e2eb6186a'
ts.set_token(token)
pro = ts.pro_api()
# 读取代码
'''
1.读取4个schema
2. 形成标准结果
3.读取估值
4.读取股价
'''

# 固定的表头
# A->M 13列
# 3行
class JunLinPlan():
    def __init__(self,sheet_names=None):
        if sheet_names is None:
            self.sheet_names = ['5G科技','自主可控','医疗健康','周期消费']
        else:
            self.sheet_names = sheet_names
        self.path =  r'C:\Users\scott\Desktop\invest\君临计划.xlsx'


        # 居中
        self._aligmentCenter = Alignment(horizontal='center', vertical='center')

        # 表题格式：黑体+14+加粗
        self._font_title = Font(u'黑体', size=11, bold=True, italic=False, strike=False, color='000000')
        self._title_width = 10.13
        # 正文格式：微软雅黑+10
        self._font_text = Font(u'微软雅黑', size=10, bold=False, italic=False, strike=False, color='000000')

        # 合理下：深蓝
        self._down_font = Font(u'微软雅黑', size=12, bold=True, italic=False, strike=False, color='191970')
        # 低估上：土耳其玉色
        self._mid1_font = Font(u'微软雅黑', size=10, bold=False, italic=False, strike=False, color='00C78C')

        # 合理上：珊瑚色
        self._mid2_font = Font(u'微软雅黑', size=10, bold=False, italic=False, strike=False, color='FF7F50')

        # 高估下：印度红
        self._up_font = Font(u'微软雅黑', size=12, bold=True, italic=False, strike=False, color='B0171F')

    def read_code(self,path,sheet_name):
        datas = pd.read_excel(path,sheet_name=sheet_name,skiprows=1,index_col=1)
        codes = datas.index
        return datas,codes


    '''
    不同的schema的股价下载
    结果规范
    index=代码
    colunms=日期
    '''
    # 下载代码
    def download_price(self,code,startday):
        """下载某一合约的分钟线数据"""
        print(f"开始下载合约数据{code}")
        df = pro.daily_basic(ts_code=code, start_date=startday)
        df.sort_values(by='trade_date', ascending=True, inplace=True)
        new_total_share = df.loc[0,'total_share'] # 最新的股本数
        df_01 = df[['trade_date','close']].swapaxes(0,1)
        # df_01.index = [code,code]
        df_01.columns = df_01.iloc[0,:]
        result = df_01.iloc[1,:]
        result.index.name = code
        result.name = code
        # print(result)
        return result,new_total_share
# 写入股价
    '''
    1.追加新的股价
    2.对比估值变色
    '''

    def writer_data(self,sheet,old_data,import_data,new_share):
        # sheet = workbook[sheetname]

        # 基础数据信息
        base_columns = 14
        base_rows = 2

        # 旧数据
        print(old_data.iloc[:,:11])
        base_total_value = old_data.iloc[:,1:5]
        print(base_total_value)
        base_stock_price = old_data.iloc[:,7:11]
        print(base_stock_price)
        data_flag = old_data.columns[-1]

        # 导入的新数据
        import_date= import_data.columns


        exist_date = []
        if re.match('\d{8}',data_flag):
            print('数据的加入')
            old_data_length = len(old_data.columns) + 1
            exist_date = old_data.columns[12:]
        else:
            print('只有基础数据')
            old_data_length = base_columns
        print('旧数据的长度:%d'%old_data_length)

        new_date = [date for  date in import_date if date not in exist_date]
        print(new_date)
        new_data = import_data.loc[:,new_date]
        print(new_data)
        if new_data.empty:
            print('没有新数据')
            return None
        new_data_length = len(new_data.columns)
        print('新数据的长度:%d'%new_data_length)

        max_row = len(old_data.index)+base_rows
        max_columns = old_data_length + new_data_length
        print('总数据的长度:%d'%max_columns)

        try:
            if old_data_length == base_columns:
                print('合并新数据')
                sheet.merge_cells(start_row=1, start_column=base_columns + 1, end_row=1, end_column=max_columns)

            else:
                print('附加新数据，先解除旧数据合并')
                sheet.unmerge_cells(start_row=1, start_column=base_columns + 1, end_row=1, end_column=old_data_length)
                sheet.merge_cells(start_row=1, start_column=base_columns+1, end_row=1, end_column=max_columns)
                # 收盘价合并
                print('合并成功')
        except ValueError:
            print('合并失败')
        sheet['N1'].alignment = Alignment(horizontal='center', vertical='center')

        print('新数据从%d列开始'%(old_data_length+1))
        for row in sheet.iter_rows(min_row=base_rows, min_col=old_data_length + 1, max_col=max_columns,
                                   max_row=max_row):
            num_row = row[0].row
            code_value = sheet.cell(num_row, column=2).value
            share_value = sheet.cell(num_row,column=8).value
            company_name = sheet.cell(num_row,column=1).value
            i = 0
            if code_value == '代码':
                for cell in row:
                    cell.value = new_data.columns[i]
                    cell.alignment = self._aligmentCenter
                    cell.font = self._font_title
                    sheet.column_dimensions[cell.column_letter].width  = self._title_width
                    i = i + 1
            else:
                print('写入%s的数据'%code_value)
                # 002405.SZ
                share_change = True
                for cell in row:
                    cell.value = new_data.loc[code_value,new_data.columns[i]]
                    if share_value != new_share[code_value] and share_change :
                        print('%s的股本发生变化，由%.2f变成%.2f'%(code_value,share_value,new_share[code_value]))
                        sheet.cell(num_row,column=8).value = new_share[code_value]
                        share_change = False
                    # 获取基础股价信息，openpyxl只能获取公式
                        down_value,mid1_value,mid2_value,up_value =[i/new_share[code_value] for i in base_total_value.loc[code_value,:]]
                    else:
                        down_value,mid1_value,mid2_value,up_value = base_stock_price.loc[code_value,:]
                    if cell.value < down_value:
                        cell.font = self._down_font
                    elif cell.value < mid1_value:
                        cell.font = self._mid1_font
                    elif cell.value > up_value:
                        cell.font = self._up_font
                    elif cell.value > mid2_value:
                        cell.font = self._mid2_font
                    else:
                        cell.font = self._font_text

                    cell.alignment = self._aligmentCenter
                    cell.comment = Comment(
                        "%s:\n"
                        "低估上：%.2f\n"
                        "合理下：%.2f\n"
                        "合理上：%.2f\n"
                        "高估下：%.2f"%(company_name,down_value,mid1_value,mid2_value,up_value ),
                        'SYW',width=120,height=100
                    )

                    i = i + 1

    def run(self,path=None,data=None):
        workbook = openpyxl.load_workbook(path)
        for sheet_name in self.sheet_names:
            print('<<<<<<开始处理:%s>>>>>>'%sheet_name)
            old_data, codes = self.read_code(self.path,sheet_name)
            import_data = None
            new_share = {}
            for code in codes:
                close_price,share = self.download_price(code, data)
                new_share[code] = float('%0.2f'%(share/10000))
                if import_data is None:
                    import_data = close_price
                else:
                    import_data = pd.concat([import_data, close_price], axis=1, sort=False)
                # print(import_data)
            import_data = import_data.swapaxes(0, 1)
            # print(import_data)
            self.writer_data(workbook[sheet_name],old_data,import_data,new_share)
        try:
            workbook.save(path)
        except PermissionError:
            print('把 “%s”关掉后，重新运行' % path)
            exit(1)



path = r'C:\Users\scott\Desktop\invest\君临计划.xlsx'
today = dt.datetime.now().strftime('%Y%m%d')
#['5G科技','自主可控','医疗健康','周期消费']
JL = JunLinPlan(['5G科技','自主可控','医疗健康'])
# JL = JunLinPlan()
# JL.writer_data()

# log_date = today
# print(today)
today = '20200717'
JL.run(path,today)











