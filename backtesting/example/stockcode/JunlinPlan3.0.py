from datetime import datetime

import openpyxl as openpyxl
import pandas as pd
from openpyxl.styles import Alignment,Font
from zsxq.database.base import get_data


# 读取代码
'''
1.读取4个schema
2. 形成标准结果
3.读取估值
4.读取股价
'''



# 固定的表头
# A->M 13列
# 3行
class JunLinPlan():

    def __init__(self,sheet_names=None):
        if sheet_names is None:
            self.sheet_names = ['5G科技','自主可控','医疗健康','周期消费']
        else:
            self.sheet_names = sheet_names
        self._begin_cell= (3,3) # C3
        self._trade_date =(2,9) # I2

        self.all_code = {}
        self.info_status = 0
        # self._max_col = 14

        # 居中
        self._aligmentCenter = Alignment(horizontal='center', vertical='center')

        # 表题格式：黑体+14+加粗
        self._font_title = Font(u'黑体', size=11, bold=True, italic=False, strike=False, color='000000')
        self._title_width = 10.13
        # 正文格式：微软雅黑+10
        self._font_text = Font(u'微软雅黑', size=10, bold=False, italic=False, strike=False, color='000000')

        # 合理下：深蓝
        self._down_font = Font(u'微软雅黑', size=12, bold=True, italic=False, strike=False, color='191970')
        # 低估上：土耳其玉色
        self._mid1_font = Font(u'微软雅黑', size=10, bold=False, italic=False, strike=False, color='00C78C')

        # 合理上：珊瑚色
        self._mid2_font = Font(u'微软雅黑', size=10, bold=False, italic=False, strike=False, color='FF7F50')

        # 高估下：印度红
        self._up_font = Font(u'微软雅黑', size=12, bold=True, italic=False, strike=False, color='B0171F')

    def read_baseinfo(self,path,sheet):
        # 原数据
        df0 = pd.read_excel(path,sheet_name=sheet,header=1,dtype={'日期':object,'交易日':object})
        info = {}
        # 公司名称
        info['name'] = df0['公司'].iloc[0]
        # 股票代码
        info['code'] = df0['代码'].iloc[0]

        # 估值基本信息
        df2 = df0.iloc[:,2:8]
        df2.columns = ['trade_date','value_down','value_mid1','value_mid2','value_up','total_share']
        df2.index = pd.to_datetime(df2['trade_date'],format='%Y%m%d')
        df2 = df2.dropna()
        del df2['trade_date']
        startday = df2.index[0]

        """  """
        # 股价的信息
        # info_status:
        df2_2 = df0.iloc[:,8:]
        df2_2 = df2_2.dropna()
        if df2_2.empty is False:
            df2_2.columns = ['trade_date', 'price_down', 'price_mid1', 'price_mid2', 'price_up', 'close']
            df2_2.index = pd.to_datetime(df2_2['trade_date'], format='%Y%m%d')
            info['last_date'] = df2_2.index[-1]
            info['old_data_len'] = len(df2_2)
        return info,df2,startday

    def deal_data(self,base_info,new_info,info):
        info['old_share'] = base_info['total_share'][-1]
        if info['old_share'] != info['new_share'] and info.get('last_date',None):
            last_date = pd.DataFrame({'total_share':info['new_share']},index=[info['last_date']])
            base_info = base_info.append(last_date)
            base_info = base_info.fillna(method='ffill')
        df = pd.concat([base_info, new_info], axis=1)
        df = df.fillna(method='ffill')
        df['price_down'] = round(df['value_down']/df['total_share'],2)
        df['price_mid1'] = round(df['value_mid1']/df['total_share'],2)
        df['price_mid2'] = round(df['value_mid2']/df['total_share'],2)
        df['price_up'] = round(df['value_up']/df['total_share'],2)
        final = df[['price_down','price_mid1','price_mid2','price_up','close']]

        return base_info,final


    '''
    不同的schema的股价下载
    结果规范
    index=代码
    colunms=日期
    '''
    # 下载代码
    def get_price(self,code,startday):
        """下载某一合约的分钟线数据"""
        print(f"开始获取合约数据{code}")
        df = get_data(ts_code=code[:-3], start_date=startday)
        new_total_share = round(df.iloc[-1,-5]/10000,2) # 最新的股本数
        result = df[['close']]
        return (result,new_total_share)
# 写入股价
    '''
    1.追加新的股价
    2.对比估值变色
    '''

    def writer_data(self,sheet,base_info,new_info,info):

        # 基础数据信息
        company_name = info['name']
        # 导入的新数据
        print("导入%d个数据"%len(new_info))
        # 判断股数有无变化
        if info['old_share'] != info['new_share']:
            print('%s的股本发生变化，由%.2f变成%.2f' % (company_name, info['old_share'],info['new_share']))

        i = 0
        for row in sheet.iter_rows(min_row=self._begin_cell[0],
                                   min_col=self._begin_cell[1],
                                   max_row=len(new_info)+2
                                   ):
            if i < len(base_info):
                base_data = base_info.iloc[i, :]
                # print((data_01.name).strftime('%Y%m%d'))
                row[0].value = base_data.name.strftime('%Y%m%d')
                row[1].value = base_data[0]
                row[2].value = base_data[1]
                row[3].value = base_data[2]
                row[4].value = base_data[3]
                row[5].value = base_data[4]

                row[0].font = self._font_text
                row[1].font = self._font_text
                row[2].font = self._font_text
                row[3].font = self._font_text
                row[4].font = self._font_text
                row[5].font = self._font_text

            # 导入交易信息
            new_data = new_info.iloc[i,:]
            # 交易日
            self.cell_format(row[6],new_data.name.strftime('%Y%m%d'),'title')

            # 低估上
            self.cell_format(row[7],new_data[0])
            # print(row[7].value)
            # 合理下
            self.cell_format(row[8],new_data[1])
            # print(row[8].value)
            # 合理上
            self.cell_format(row[9],new_data[2])
            # print(row[9].value)
            # 高估下
            self.cell_format(row[10],new_data[3])
            # print(row[10].value)
            # 收盘价
            self.cell_format(row[11],new_data[4])
            # print(row[11].value)
            if row[11].value < row[7].value:
                row[11].font = self._down_font
            elif row[11].value < row[8].value:
                row[11].font = self._mid1_font
            elif row[11].value > row[10].value:
                row[11].font = self._up_font
            elif row[11].value > row[9].value:
                row[11].font = self._mid2_font
            else:
                row[11].font = self._font_text
            i = i+1

    def cell_format(self,cell,value,type=None):
        cell.value = value
        cell.alignment = self._aligmentCenter
        if type == 'title':
            cell.font = self._font_title
        elif type is None:
            cell.font = self._font_text

    def run(self,path=None):
        workbook = openpyxl.load_workbook(path)
        all_name = workbook.sheetnames
        for sheet_name in all_name:
            print('<<<<<<开始处理:%s>>>>>>'%sheet_name)
            sheet = workbook[sheet_name]
            info,base_info,download_date = self.read_baseinfo(path,sheet_name)
            new_info,total_share = self.get_price(info['code'], download_date)
            if info.get('old_data_len',None) and len(new_info) == info['old_data_len']:
                print('[WARNNING]<<<<<<%s没有新数据导入>>>>>>' % sheet_name)
                continue
            info['new_share'] = total_share
            try:
                base_info,final = self.deal_data(base_info,new_info,info)
            except ValueError:
                print('【ERROR】<<<<<<%s导入失败>>>>>>'%sheet_name)
                continue
            self.writer_data(sheet,base_info,final,info)

            try:
                print('<%s>数据更新完毕'%info['name'])
                workbook.save(path)
            except PermissionError:
                print('把 “%s”关掉后，重新运行' % path)
                exit(1)



# path1 = r'C:\Users\scott\Desktop\invest\君临计划-消费.xlsx'
# path2 = r'C:\Users\scott\Desktop\invest\君临计划-科技.xlsx'
# path3 = r'C:\Users\scott\Desktop\invest\君临计划-医药.xlsx'
path1 = r'X:\股票\君临计划-消费.xlsx'
path2 = r'X:\股票\君临计划-科技.xlsx'
path3 = r'X:\股票\君临计划-医药.xlsx'
paths = [
    path1,
    path2,
    path3
]
JL = JunLinPlan()
for p in paths :
    JL.run(p)
