
import openpyxl as openpyxl
import tushare as ts
from openpyxl.styles import Alignment,Font

token = 'bfbf67e56f47ef62e570fc6595d57909f9fc516d3749458e2eb6186a'
ts.set_token(token)
pro = ts.pro_api()
# 读取代码
'''
1.读取4个schema
2. 形成标准结果
3.读取估值
4.读取股价
'''



# 固定的表头
# A->M 13列
# 3行
class JunLinPlan():
    def __init__(self,sheet_names=None):
        if sheet_names is None:
            self.sheet_names = ['5G科技','自主可控','医疗健康','周期消费']
        else:
            self.sheet_names = sheet_names
        self._base_date= (2,3) # D2
        self._trade_date =(2,9) # I2
        self.all_code = {}

        # self._max_col = 14

        # 居中
        self._aligmentCenter = Alignment(horizontal='center', vertical='center')

        # 表题格式：黑体+14+加粗
        self._font_title = Font(u'黑体', size=11, bold=True, italic=False, strike=False, color='000000')
        self._title_width = 10.13
        # 正文格式：微软雅黑+10
        self._font_text = Font(u'微软雅黑', size=10, bold=False, italic=False, strike=False, color='000000')

        # 合理下：深蓝
        self._down_font = Font(u'微软雅黑', size=12, bold=True, italic=False, strike=False, color='191970')
        # 低估上：土耳其玉色
        self._mid1_font = Font(u'微软雅黑', size=10, bold=False, italic=False, strike=False, color='00C78C')

        # 合理上：珊瑚色
        self._mid2_font = Font(u'微软雅黑', size=10, bold=False, italic=False, strike=False, color='FF7F50')

        # 高估下：印度红
        self._up_font = Font(u'微软雅黑', size=12, bold=True, italic=False, strike=False, color='B0171F')

    def read_baseinfo(self,sheet,beg_date):
        name = (3,1) # A3
        code = (3,2) # B3
        maxrow = sheet.max_row
        name = sheet.cell(row=name[0],column=name[1])
        code = sheet.cell(row=code[0],column=code[1])

        result = {}
        base_start_row = 0
        trade_start_row = 0
        for i in range(self._trade_date[0]+1,maxrow+1):
            basedate = sheet.cell(i,self._base_date[1]).value
            tradedate = sheet.cell(i,self._trade_date[1]).value
            if basedate is None and result.get('base_start_row',None) is None:
                base_start_row = i-1
                result['base_start_row'] = base_start_row
            if tradedate is None and trade_start_row == 0 :
                trade_start_row = i
            if self._trade_date is not None and trade_start_row == 0:
                if int(beg_date) <= int(tradedate):
                    trade_start_row = i
        if trade_start_row == 0:
            trade_start_row = i + 1

        # 市值估值的读取
        result['trade_start_row'] = trade_start_row
        result['base_start_row'] = base_start_row
        result['total_share'] = sheet.cell(base_start_row,self._base_date[1]+5).value
        result['name'] = name.value
        result['code'] = code.value
        result['sheet_maxrow'] = maxrow
        self.all_code[name.value] = code.value
        print('<%s>生成基础数据'%result['name'])

        last_evalue_date = sheet.cell(base_start_row,self._base_date[1]).value
        # 判断下载数据的日期与最新估值的日期，不在更新旧数据和新估值的关系了
        if int(last_evalue_date) >= int(beg_date) :
            download_date = last_evalue_date
        else:
            download_date = beg_date
        return result,download_date

    '''
    不同的schema的股价下载
    结果规范
    index=代码
    colunms=日期
    '''
    # 下载代码
    def download_price(self,code,startday):
        """下载某一合约的分钟线数据"""
        print(f"开始下载合约数据{code}")
        df = pro.daily_basic(ts_code=code, start_date=startday)
        df.sort_values(by='trade_date', ascending=True, inplace=True)
        new_total_share = round(df.loc[0,'total_share']/10000,2) # 最新的股本数
        result = df[['close']]
        result.index = df['trade_date']
        # result.to_csv('603345_01.csv',header=None)
        return (result,new_total_share)
# 写入股价
    '''
    1.追加新的股价
    2.对比估值变色
    '''

    def writer_data(self,sheet,base_info,new_info,date):

        # 基础数据信息
        company_name = base_info['name']
        last_trade_row = base_info['trade_start_row']
        base_rows =base_info['base_start_row']
        undervalue_up = sheet.cell(base_rows,self._base_date[1]+1).value
        reason_down = sheet.cell(base_rows,self._base_date[1]+2).value
        reason_up = sheet.cell(base_rows,self._base_date[1]+3).value
        overvalue_down = sheet.cell(base_rows,self._base_date[1]+4).value


        # 导入的新数据
        new_share = new_info[1]
        new_date = date
        new_data = new_info[0]['close']
        print("导入%d个数据"%len(new_data))

        # 判断股数有无变化
        old_share = base_info['total_share']
        if old_share != new_share:
            print('%s的股本发生变化，由%.2f变成%.2f' % (company_name, old_share, new_share))
            base_write_flag = True
        else:
            base_write_flag = False
        total_shares = new_share

        # 股本写入
        if base_write_flag:
            self.cell_format(sheet.cell(base_rows+1,self._base_date[1]),new_date)
            self.cell_format(sheet.cell(base_rows+1,self._base_date[1]+1),undervalue_up)
            self.cell_format(sheet.cell(base_rows+1,self._base_date[1]+2),reason_down)
            self.cell_format(sheet.cell(base_rows+1,self._base_date[1]+3),reason_up)
            self.cell_format(sheet.cell(base_rows+1,self._base_date[1]+4),overvalue_down)
            self.cell_format(sheet.cell(base_rows+1,self._base_date[1]+5),total_shares)

        down_value = round(undervalue_up / total_shares,2)
        mid1_value = round(reason_down / total_shares,2)
        mid2_value = round(reason_up / total_shares,2)
        up_value   = round(overvalue_down / total_shares,2)

        max_row = last_trade_row + len(new_data)
        if max_row <= base_info['sheet_maxrow']:
            max_row = base_info['sheet_maxrow']
        max_columns = self._trade_date[1] + 5
        print('数据总的行数:%d'%max_row)
        print('写入<%s>的数据,数据从%s开始' % (company_name,date))
        i = 0
        for row in sheet.iter_rows(min_row=last_trade_row, min_col=self._trade_date[1], max_col=max_columns,
                                   max_row=max_row):
            if i >= len(new_data):
                row[0].value = None
                row[1].value = None
                row[2].value = None
                row[3].value = None
                row[4].value = None
                row[5].value = None
            else:
                # 交易日
                self.cell_format(row[0],new_data.index[i],'title')

                # 低估上
                self.cell_format(row[1],down_value)
                # print(row[1].value)
                # 合理下
                self.cell_format(row[2],mid1_value)
                # print(row[2].value)
                # 合理上
                self.cell_format(row[3],mid2_value)
                # print(row[3].value)
                # 高估下
                self.cell_format(row[4],up_value)
                # print(row[4].value)
                # 收盘价
                self.cell_format(row[5],new_data.iloc[i])
                # print(row[5].value)
                if row[5].value < down_value:
                    row[5].font = self._down_font
                elif row[5].value < mid1_value:
                    row[5].font = self._mid1_font
                elif row[5].value > up_value:
                    row[5].font = self._up_font
                elif row[5].value > mid2_value:
                    row[5].font = self._mid2_font
                else:
                    row[5].font = self._font_text
            i = i+1

    def cell_format(self,cell,value,type=None):
        cell.value = value
        cell.alignment = self._aligmentCenter
        if type == 'title':
            cell.font = self._font_title
        elif type is None:
            cell.font = self._font_text

    def run(self,path=None,date=None):
        workbook = openpyxl.load_workbook(path)
        all_name = workbook.sheetnames
        # all_name.remove('安井食品')
        # for sheet_name in self.sheet_names:
        for sheet_name in all_name:
            print('<<<<<<开始处理:%s>>>>>>'%sheet_name)
            sheet = workbook[sheet_name]
            base_info,download_date = self.read_baseinfo(sheet,date)
            # new_info = (pd.read_csv('603345_01.csv', names=['close_price'], header=1, index_col=0),3)
            # value = new_info.iloc[1]
            # print(value.values)
            new_info = self.download_price(base_info['code'], download_date)

            self.writer_data(sheet,base_info,new_info,download_date)

            try:
                print('<%s>数据更新完毕'%base_info['name'])
                workbook.save(path)
            except PermissionError:
                print('把 “%s”关掉后，重新运行' % path)
                exit(1)



path1 = r'C:\Users\scott\Desktop\invest\君临计划-消费.xlsx'
path2 = r'C:\Users\scott\Desktop\invest\君临计划-科技.xlsx'
path3 = r'C:\Users\scott\Desktop\invest\君临计划-医药.xlsx'
paths = [path1,path2,path3]
JL = JunLinPlan()
today = '20200920'
for p in paths :
    JL.run(p,today)


print(JL.all_code)








