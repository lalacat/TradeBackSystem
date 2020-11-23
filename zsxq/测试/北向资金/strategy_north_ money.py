#!/usr/bin/env python
# coding: utf-8
import os
from datetime import datetime

import openpyxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pylab import mpl
import tushare as ts
import time

# # 策略实例

# 基于北向资金变动数据构建布林带择时策略能有效判断市场短期涨跌
#
# 至于北向资金是否具有指示作用，我们可以通过择时模型的回测表现来进行观测。为此， 我们基于北向资金的流入规模数据构建布林带策略：
#
# （1） 当该日北向资金流入规模 > 过去 252 个交易日的北向资金均值 + 1.5 倍标准差， 则全仓买入沪深 300；
#
# （2） 当该日北向资金流入规模 < 过去 252 个交易日的北向资金均值 - 1.5 倍标准差， 则清仓卖出沪深 300；
#
# （3） 以第二天开盘价买入计量。

# In[102]:
# from zsxq.测试.北向资金.northmoney import get_index_data, get_north_money, indexs


class North_Strategy(object):

    def __init__(self):
        token = 'e0eeb08befd1f07516df2cbf9cbd58663f77fd72f92a04f290291c9d'
        self.pro = ts.pro_api(token)


    def init_parm(self,data, window, stdev_n, cost):
        '''输入参数：
        data:[收盘价close,开盘价open,北上资金north]
        window:移动窗口
        stdev_n:几倍标准差
        cost:手续费
        '''
        # 中轨
        df = data.copy().dropna()
        df['mid'] = df['北向资金'].rolling(window).mean()
        stdev = df['北向资金'].rolling(window).std()
        # 上下轨
        df['upper'] = df['mid'] + stdev_n * stdev
        df['lower'] = df['mid'] - stdev_n * stdev
        df['ret'] = df.close / df.close.shift(1) - 1
        df.dropna(inplace=True)

        # 设计买卖信号
        # 当日北向资金突破上轨线发出买入信号设置为1,下个交易日开仓
        df.loc[df['北向资金'] > df.upper, 'signal'] = 1
        # 当日北向资金跌破下轨线发出卖出信号设置为0，下个交易日平仓
        df.loc[df['北向资金'] < df.lower, 'signal'] = 0
        print(df['signal'])
        df['position'] = df['signal'].shift(1)
        df['position'].fillna(method='ffill', inplace=True)
        df['position'].fillna(0, inplace=True)
        # 根据交易信号和仓位计算策略的每日收益率
        df.loc[df.index[0], 'capital_ret'] = 0
        # 今天开盘新买入的position在今天的涨幅(扣除手续费)
        df.loc[df['position'] > df['position'].shift(1), 'capital_ret'] = (df.close / df.open - 1) * (1 - cost)
        # 卖出同理
        df.loc[df['position'] < df['position'].shift(1), 'capital_ret'] = (df.open / df.close.shift(1) - 1) * (1 - cost)
        # 当仓位不变时,当天的capital是当天的change * position
        df.loc[df['position'] == df['position'].shift(1), 'capital_ret'] = df['ret'] * df['position']
        print(df)
        # 计算标的、策略、指数的累计收益率
        df['策略净值'] = (df.capital_ret + 1.0).cumprod()
        df['指数净值'] = (df.ret + 1.0).cumprod()
        return df

    def download_data(self,code,start,end=None):
        print('%s正在下载' % code)
        stock_data = self.get_code_data(code, start, end)
        north_data = self.get_north_money(start, end)
        result_df = stock_data.join(north_data['north_money'], how='inner')
        result_df.rename(columns={'north_money': '北向资金'}, inplace=True)
        result_df = result_df[['close', 'open', '北向资金']].dropna()
        return result_df


    # 策略评价指标
    def performance(self,df):
        df1 = df.loc[:, ['ret', 'capital_ret']]
        # 计算每一年(月,周)股票,资金曲线的收益
        year_ret = df1.resample('A').apply(lambda x: (x + 1.0).prod() - 1.0)
        month_ret = df1.resample('M').apply(lambda x: (x + 1.0).prod() - 1.0)
        week_ret = df1.resample('W').apply(lambda x: (x + 1.0).prod() - 1.0)
        # 去掉缺失值
        year_ret.dropna(inplace=True)
        month_ret.dropna(inplace=True)
        week_ret.dropna(inplace=True)
        # 计算策略的年（月，周）胜率
        year_win_rate = len(year_ret[year_ret['capital_ret'] > 0]) / len(year_ret[year_ret['capital_ret'] != 0])
        month_win_rate = len(month_ret[month_ret['capital_ret'] > 0]) / len(month_ret[month_ret['capital_ret'] != 0])
        week_win_rate = len(week_ret[week_ret['capital_ret'] > 0]) / len(week_ret[week_ret['capital_ret'] != 0])
        # 计算总收益率、年化收益率和风险指标
        total_ret = df[['策略净值', '指数净值']].iloc[-1] - 1
        annual_ret = pow(1 + total_ret, 250 / len(df1)) - 1
        dd = (df[['策略净值', '指数净值']].cummax() - df[['策略净值', '指数净值']]) / df[['策略净值', '指数净值']].cummax()
        d = dd.max()
        beta = df[['capital_ret', 'ret']].cov().iat[0, 1] / df['ret'].var()
        alpha = (annual_ret['策略净值'] - annual_ret['指数净值'] * beta)
        exReturn = df['capital_ret'] - 0.03 / 250
        sharper_atio = np.sqrt(len(exReturn)) * exReturn.mean() / exReturn.std()
        TA1 = round(total_ret['策略净值'] * 100, 2)
        TA2 = round(total_ret['指数净值'] * 100, 2)
        AR1 = round(annual_ret['策略净值'] * 100, 2)
        AR2 = round(annual_ret['指数净值'] * 100, 2)
        MD1 = round(d['策略净值'] * 100, 2)
        MD2 = round(d['指数净值'] * 100, 2)
        S = round(sharper_atio, 2)
        # 输出结果
        print(f'策略年胜率为：{round(year_win_rate * 100, 2)}%')
        print(f'策略月胜率为：{round(month_win_rate * 100, 2)}%')
        print(f'策略周胜率为：{round(week_win_rate * 100, 2)}%')

        print(f'总收益率：  策略：{TA1}%，沪深300：{TA2}%')
        print(f'年化收益率：策略：{AR1}%, 沪深300：{AR2}%')
        print(f'最大回撤：  策略：{MD1}%, 沪深300：{MD2}%')
        print(f'策略Alpha： {round(alpha, 2)}, Beta：{round(beta, 2)}，夏普比率：{S}')

    # 对策略累计收益率进行可视化
    def plot_performance(self,df, name):
        d1 = df[['策略净值', '指数净值', 'signal']]
        d1[['策略净值', '指数净值']].plot(figsize=(15, 7))

        for i in d1.index:
            v = d1['指数净值'][i]
            if d1.signal[i] == 1:
                plt.scatter(i, v, c='r')
            if d1.signal[i] == 0:
                plt.scatter(i, v, c='g')

        plt.title(name + '—' + '北向资金择时交易策略回测', size=15)
        plt.xlabel('')
        ax = plt.gca()
        ax.spines['right'].set_color('none')
        ax.spines['top'].set_color('none')
        plt.show()

    def _save_data(self, code, data, path):
        try:
            if not os.path.exists(path):
                wb = openpyxl.Workbook()
                wb.remove(wb['Sheet'])
            else:
                wb = openpyxl.load_workbook(path)
            all_sheet = wb.sheetnames
            if all_sheet:
                if code in all_sheet:
                    wb.remove(wb[code])
            wb.create_sheet(code)
            sheet = wb[code]
            # 将dataframe类型的数据转换为list
            # 先转换为narray,要对index数据进行字符串化
            # 在转换为list类型的数据
            new_datas = data.reset_index()
            index = new_datas['trade_date'].apply(lambda x: datetime.strftime(x, '%Y-%m-%d'))
            new_datas['trade_date'] = index
            sheet.append(new_datas.columns.values.tolist())
            for i in new_datas.values:
                sheet.append(i.tolist())
            wb.save(path)
            print('%s写入完成' % code)
        except Exception as e:
            print(e)

    def get_from_excel(self, code, path):
        df = pd.read_excel(path, header=0, index_col=0, sheet_name=[code])[code]
        df.index = pd.to_datetime(df.index)
        return df

        # 获取交易日历

    def _get_cal_date(self, start, end=None):
        cal_date = self.pro.trade_cal(exchange='', start_date=start, end_date=end)
        cal_date = cal_date[cal_date.is_open == 1]
        dates = cal_date.cal_date.values
        return dates

        # 获取指数数据

    def get_code_data(self, code, start, end):
        index_df = self.pro.index_daily(ts_code=code, start_date=start, end_date=end)
        index_df.index = pd.to_datetime(index_df.trade_date)
        index_df = index_df.sort_index()
        return index_df

        # 获取北向资金数据

    def get_north_money(self, start, end):
        # 获取交易日历
        dates = self._get_cal_date(start, end)
        # tushare限制流量，每次只能获取300条记录
        df = self.pro.moneyflow_hsgt(start_date=start, end_date=end)
        if len(df) <= 300:
            df = df.drop_duplicates()
            df.index = pd.to_datetime(df.trade_date)
            df = df.sort_index()
        else:
            # 拆分时间进行拼接，再删除重复项
            for i in range(0, len(dates) - 300, 300):
                d0 = self.pro.moneyflow_hsgt(start_date=dates[i], end_date=dates[i + 300])
                df = pd.concat([d0, df])
                # 删除重复项
                df = df.drop_duplicates()
                df.index = pd.to_datetime(df.trade_date)
                df = df.sort_index()
        return df


#将上述函数整合成一个执行函数
# def main(code='000300.SH',start='20141117',end='20200812',window=252,stdev_n=1.5,cost=0.00):
#     hs300=get_index_data(code,start,end)
#     print(hs300)
#     north_data=get_north_money(start,end)
#     print(north_data)
#     result_df=hs300.join(north_data['north_money'],how='inner')
#     result_df.rename(columns={'north_money':'北向资金'},inplace=True)
#     result_df=result_df[['close','open','北向资金']].dropna()
#     print(result_df)
#     df=North_Strategy(result_df,window,stdev_n,cost)
#     name=list (indexs.keys()) [list (indexs.values()).index (code)]
#     print(f'回测标的：{name}指数')
#     startDate=df.index[0].strftime('%Y%m%d')
#     print(f'回测期间：{startDate}—{end}')
#     performance(df)
#     plot_performance(df,name)


#

start='20190101'
# 当前时间
now = datetime.strftime(datetime.now(),'%Y%m%d')
code_1 = '000300.SH'
code_2 = '000905.SH'
index_path = 'X:\\股票\\indexs.xlsx'
result_path = 'X:\\股票\\result.xlsx'

# main('000300.SH',start,now,22,1.5,0.01)
nm = North_Strategy()
# data = nm.download_data(code_2,start,now)
# nm._save_data(code_2,data,index_path)
data = nm.get_from_excel(code_2,index_path)
result = nm.init_parm(data,20,1.65,0.01)
# nm._save_data(code_2,result,result_path)